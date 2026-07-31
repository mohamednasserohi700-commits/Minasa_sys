"""Report generation: Excel (openpyxl) and PDF (reportlab)."""
import io
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def requests_to_excel(requests) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Requests"

    headers = ["Ticket #", "Full Name", "Company", "Email", "Phone", "Country", "City",
               "Project Name", "Platform", "Budget", "Status", "Submitted On"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in requests:
        ws.append([
            r.ticket_number, r.full_name, r.company or "", r.email, r.phone,
            r.country or "", r.city or "", r.project_name, r.platform or "",
            r.budget or "", r.status,
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 12), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def requests_to_pdf(requests, title="Project Requests Report") -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleStyle", parent=styles["Title"], textColor=colors.HexColor("#4F46E5"))
    meta_style = ParagraphStyle("Meta", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    elements = [
        Paragraph(title, title_style),
        Paragraph(f"Generated on {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} &middot; "
                  f"Total records: {len(requests)}", meta_style),
        Spacer(1, 12),
    ]

    data = [["Ticket #", "Client", "Project", "Platform", "Status", "Submitted"]]
    for r in requests:
        data.append([
            r.ticket_number, r.full_name[:22], r.project_name[:26], r.platform or "-",
            r.status, r.created_at.strftime("%Y-%m-%d") if r.created_at else "-",
        ])

    table = Table(data, repeatRows=1, colWidths=[70, 95, 120, 75, 70, 65])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FE")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def single_request_to_pdf(r) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleStyle", parent=styles["Title"], textColor=colors.HexColor("#4F46E5"))
    h_style = ParagraphStyle("H", parent=styles["Heading3"], textColor=colors.HexColor("#4F46E5"), spaceBefore=10)
    body_style = ParagraphStyle("Body", parent=styles["Normal"], spaceAfter=6)

    elements = [Paragraph(f"Ticket {r.ticket_number}", title_style),
                Paragraph(f"Status: <b>{r.status}</b>", body_style), Spacer(1, 10)]

    def field(label, value):
        return Paragraph(f"<b>{label}:</b> {value or '-'}", body_style)

    elements.append(Paragraph("Client Information", h_style))
    for label, value in [
        ("Full Name", r.full_name), ("Company", r.company), ("Email", r.email),
        ("Phone", r.phone), ("WhatsApp", r.whatsapp), ("Country", r.country),
        ("City", r.city), ("Business Type", r.business_type),
    ]:
        elements.append(field(label, value))

    elements.append(Paragraph("Project Details", h_style))
    for label, value in [
        ("Project Name", r.project_name), ("Category", r.project_category),
        ("Platform", r.platform), ("Description", r.description),
        ("Expected Features", r.expected_features), ("Target Users", r.target_users),
        ("Budget", r.budget), ("Delivery Time", r.delivery_time),
        ("Existing System", r.existing_system), ("Preferred Contact", r.preferred_contact),
        ("Additional Notes", r.additional_notes),
    ]:
        elements.append(field(label, value))

    doc.build(elements)
    buffer.seek(0)
    return buffer
